import streamlit as st
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import traceback

# ─── PAGE CONFIG ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Tạo SBD Tự Động",
    page_icon="💮",
    layout="centered",
    initial_sidebar_state="collapsed",
)

# ─── CSS ───────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
/* ── Force light mode toàn bộ, bỏ qua prefers-color-scheme ── */
html, body, [data-testid="stAppViewContainer"], .stApp,
[data-testid="stApp"], [class*="main"] {
    background-color: #FFFFFF !important;
    color: #1E293B !important;
}
[data-testid="stHeader"] { background: transparent !important; box-shadow: none !important; }
[data-testid="stSidebar"] { background: #F8FAFC !important; }

/* ── Ép tất cả text mặc định về màu tối ── */
p, span, label, div, h1, h2, h3, h4, li,
[data-testid="stMarkdownContainer"] p,
[data-testid="stMarkdownContainer"] span {
    color: #1E293B !important;
}

/* ── Title ── */
.app-title {
    text-align: center;
    font-size: 1.75rem;
    font-weight: 800;
    padding: 1.2rem 0 .25rem;
    color: #1D4ED8 !important;
    letter-spacing: .3px;
}
.app-subtitle {
    text-align: center;
    color: #475569 !important;
    font-size: .9rem;
    margin-bottom: 1.6rem;
}

/* ── Cards: nền sữa đậm, viền xanh nhạt ── */
.card {
    background: #EEF2F7 !important;
    border-radius: 14px;
    padding: 1.15rem 1.5rem 1.25rem;
    margin-bottom: 1rem;
    border: 1.5px solid #C7D9F0;
    box-shadow: 0 2px 8px rgba(30,58,138,.06);
}
.card-label {
    font-size: .72rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 1.3px;
    color: #2563EB !important;
    margin-bottom: .5rem;
}
.card-title {
    font-size: 1rem;
    font-weight: 700;
    color: #0F172A !important;
    margin-bottom: .8rem;
}

/* ── Radio pills ── */
div[data-testid="stRadio"] > div { gap: 8px !important; flex-wrap: wrap; }
div[data-testid="stRadio"] label {
    background: #FFFFFF !important;
    border: 1.5px solid #93C5FD !important;
    border-radius: 50px !important;
    padding: .38rem 1.1rem !important;
    font-weight: 600 !important;
    font-size: .9rem !important;
    color: #1D4ED8 !important;
    transition: all .15s;
}
div[data-testid="stRadio"] label:has(input:checked) {
    background: #1D4ED8 !important;
    border-color: #1D4ED8 !important;
    color: #FFFFFF !important;
}
div[data-testid="stRadio"] label * { color: inherit !important; }

/* ── Caption / small text ── */
[data-testid="stCaptionContainer"] p,
small { color: #475569 !important; }

/* ── Download button ── */
.stDownloadButton > button {
    background: #FFFFFF !important;
    border: 1.5px solid #93C5FD !important;
    color: #1D4ED8 !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
}
.stDownloadButton > button:hover { background: #EFF6FF !important; }
.stButton > button { border-radius: 8px !important; font-weight: 600 !important; }

/* ── Info / banners ── */
.banner-ok {
    background: #F0FDF4; border-left: 4px solid #22C55E;
    padding: .8rem 1.1rem; border-radius: 8px; margin: .7rem 0;
    color: #15803D !important; font-weight: 600; font-size: .92rem;
}
.banner-err {
    background: #FEF2F2; border-left: 4px solid #EF4444;
    padding: .8rem 1.1rem; border-radius: 8px; margin: .5rem 0;
    color: #B91C1C !important; font-weight: 600; font-size: .92rem;
}
/* Streamlit's st.info box */
[data-testid="stAlert"] { background: #EFF6FF !important; border-color: #93C5FD !important; }
[data-testid="stAlert"] p { color: #1E3A8A !important; }

/* ── File uploader ── */
[data-testid="stFileUploader"] {
    border: 2px dashed #93C5FD !important;
    border-radius: 10px !important;
    background: #F8FBFF !important;
}
[data-testid="stFileUploaderDropzone"] {
    background: #F8FBFF !important;
    color: #334155 !important;
}
[data-testid="stFileUploaderDropzone"] * { color: #334155 !important; }

/* ── Metric cards ── */
[data-testid="stMetric"] {
    background: #DBEAFE !important;
    border-radius: 10px;
    padding: .6rem .8rem;
    border: 1px solid #BFDBFE;
}
[data-testid="stMetricValue"], [data-testid="stMetricLabel"] {
    color: #1E3A8A !important;
}

/* ── Dataframe ── */
[data-testid="stDataFrame"] { border-radius: 10px; overflow: hidden; }

/* ── Step badge (cột trái) ── */
.step-badge {
    display: inline-flex;
    align-items: center;
    justify-content: center;
    background: #1D4ED8;
    color: #FFFFFF !important;
    font-size: .78rem;
    font-weight: 700;
    letter-spacing: .5px;
    border-radius: 8px;
    padding: .35rem .75rem;
    margin-top: .15rem;
    white-space: nowrap;
}

/* ── Divider giữa các bước ── */
.row-divider {
    border-top: 1.5px solid #DBEAFE;
    margin: .8rem 0;
}

/* ── Sub-label trong bước 2 ── */
.sub-label {
    font-size: .8rem;
    font-weight: 700;
    color: #2563EB !important;
    margin-bottom: .3rem;
    text-transform: uppercase;
    letter-spacing: .8px;
}

/* ── Note kỳ thi ── */
.exam-note {
    font-size: .83rem;
    color: #475569 !important;
    background: #F0F6FF;
    border-left: 3px solid #93C5FD;
    border-radius: 0 6px 6px 0;
    padding: .45rem .8rem;
    margin-top: .5rem;
}

/* ── Wrapper toàn bộ 3 bước ── */
.steps-wrapper {
    background: #FFFFFF;
    border: 1.5px solid #DBEAFE;
    border-radius: 14px;
    padding: 1.2rem 1.4rem;
    box-shadow: 0 2px 10px rgba(37,99,235,.06);
    margin-bottom: 1.2rem;
}
</style>
""", unsafe_allow_html=True)


# ─── CONSTANTS ─────────────────────────────────────────────────────────────────
EXAM_RULES = {
    "BEBRAS": {
        "label": "BEBRAS",
        "desc": "Lớp 1–12, chia 6 cấp độ",
        "valid_grades": list(range(1, 13)),
    },
    "AMC8": {
        "label": "AMC8",
        "desc": "Lớp 4–8",
        "valid_grades": [4, 5, 6, 7, 8],
    },
    "AMC1012": {
        "label": "AMC10/12",
        "desc": "AMC10: Lớp 6–10 | AMC12: Lớp 11–12",
        "valid_grades": list(range(6, 13)),
        "sub_levels": {"AMC10": [6,7,8,9,10], "AMC12": [11,12]},
    },
    "VEO": {
        "label": "VEO",
        "desc": "VEO JUNIOR: Lớp 6–9 | VEO: Lớp 10–12",
        "valid_grades": list(range(6, 13)),
        "sub_levels": {"VEO JUNIOR": [6,7,8,9], "VEO": [10,11,12]},
        "level_order": ["VEO JUNIOR", "VEO"],
    },
}

SAMPLE_PATH = "SBD_mẫu.xlsx"


# ─── HELPERS ───────────────────────────────────────────────────────────────────
def load_sample_bytes():
    with open(SAMPLE_PATH, "rb") as f:
        return f.read()


def assign_level_bebras(grade):
    for pair, lv in [((1,2),1),((3,4),2),((5,6),3),((7,8),4),((9,10),5),((11,12),6)]:
        if grade in pair: return lv
    return None


def assign_sub_level(exam, grade):
    for name, grades in EXAM_RULES[exam].get("sub_levels", {}).items():
        if grade in grades: return name
    return None


def validate_input(df_hs, df_diem, df_xep, exam):
    errors = []
    for c in ["STT","ID gốc","Họ và tên đệm","Tên","Ngày sinh","Tháng sinh",
              "Năm sinh","Giới tính","Khối lớp","Lớp","Xã/Phường","Tỉnh thành",
              "Cấp độ","HS nước ngoài","Cụm"]:
        if c not in df_hs.columns:
            errors.append(f"Sheet 'Học-sinh' thiếu cột: **{c}**")
    for c in ["Điểm thi","Phòng","Số lượng hs/phòng"]:
        if c not in df_diem.columns:
            errors.append(f"Sheet 'Điểm-thi' thiếu cột: **{c}**")
    for c in ["Cụm","Điểm thi"]:
        if c not in df_xep.columns:
            errors.append(f"Sheet 'Xếp-điểm-thi' thiếu cột: **{c}**")
    if errors:
        return errors

    valid = EXAM_RULES[exam]["valid_grades"]
    bad = df_hs[~df_hs["Khối lớp"].isin(valid)]["Khối lớp"].dropna().unique()
    if len(bad):
        errors.append(f"Có học sinh không thuộc khối hợp lệ cho {exam}: {list(bad)}")

    cum_list = df_hs["Cụm"].dropna().astype(str).str.strip().unique()
    xep_v = df_xep.dropna(subset=["Cụm","Điểm thi"]).copy()
    xep_v["Cụm"] = xep_v["Cụm"].astype(str).str.strip()
    xep_map = xep_v.set_index("Cụm")["Điểm thi"].to_dict()
    missing = [c for c in cum_list if c not in xep_map]
    if missing:
        errors.append(f"Các Cụm chưa được gán Điểm thi: {missing}")

    return errors


def process_sbd(df_hs_raw, df_diem, df_xep, exam):
    df = df_hs_raw.copy()
    df = df[df["Khối lớp"].notna() & df["Tên"].notna()].copy()
    df = df[df["Khối lớp"].isin(EXAM_RULES[exam]["valid_grades"])].copy()
    if len(df) == 0:
        raise ValueError("Không có học sinh hợp lệ sau khi lọc theo kỳ thi.")

    df["HS nước ngoài"] = pd.to_numeric(df["HS nước ngoài"], errors="coerce").fillna(0).astype(int)
    df["Khối lớp"] = pd.to_numeric(df["Khối lớp"], errors="coerce").astype(int)

    # ── Gán cấp độ ──
    if exam == "BEBRAS":
        df["Cấp độ"] = df["Khối lớp"].apply(assign_level_bebras)
    elif exam in ("AMC1012", "VEO"):
        df["Cấp độ"] = df["Khối lớp"].apply(lambda g: assign_sub_level(exam, g))
    else:
        df["Cấp độ"] = df["Cấp độ"].fillna("")

    # ── Sort Phase 1 ──
    if exam == "VEO":
        # VEO JUNIOR trước → VEO sau (dùng Categorical có thứ tự)
        level_order = EXAM_RULES["VEO"]["level_order"]
        df["Cấp độ"] = pd.Categorical(df["Cấp độ"], categories=level_order, ordered=True)
        df = df.sort_values(
            ["Cấp độ", "Khối lớp", "Tên", "Họ và tên đệm"],
            ascending=True, na_position="last"
        ).reset_index(drop=True)
        df["Cấp độ"] = df["Cấp độ"].astype(str)
    else:
        df = df.sort_values(
            ["HS nước ngoài", "Cấp độ", "Khối lớp", "Tên", "Họ và tên đệm"],
            ascending=True, na_position="last"
        ).reset_index(drop=True)

    df["STT"] = range(1, len(df) + 1)

    # ── Map Điểm thi ──
    xep_clean = df_xep.dropna(subset=["Cụm","Điểm thi"]).copy()
    xep_clean["Cụm"] = xep_clean["Cụm"].astype(str).str.strip()
    xep_map = xep_clean.set_index("Cụm")["Điểm thi"].to_dict()
    df["Cụm"] = df["Cụm"].astype(str).str.strip()
    df["Điểm thi"] = df["Cụm"].map(xep_map)

    # ── Build room capacity map ──
    dc = df_diem.dropna(subset=["Điểm thi","Phòng","Số lượng hs/phòng"]).copy()
    for col in ["Điểm thi","Phòng","Số lượng hs/phòng"]:
        dc[col] = pd.to_numeric(dc[col], errors="coerce")
    dc = dc.dropna()
    room_map = {
        dt: sorted(grp[["Phòng","Số lượng hs/phòng"]].values.tolist(), key=lambda x: x[0])
        for dt, grp in dc.groupby("Điểm thi")
    }

    # ── Gán phòng ──
    phong_col, stt_phong_col = [], []
    room_state = {}  # dt -> [room_idx, count_in_room]
    for _, row in df.iterrows():
        dt = row["Điểm thi"]
        if pd.isna(dt):
            phong_col.append(None); stt_phong_col.append(None); continue
        dt = int(dt)
        rooms = room_map.get(dt, [])
        if not rooms:
            phong_col.append(None); stt_phong_col.append(None); continue
        if dt not in room_state:
            room_state[dt] = [0, 0]
        idx, cnt = room_state[dt]
        # chuyển phòng nếu đầy
        if cnt >= int(rooms[idx][1]) and idx + 1 < len(rooms):
            idx += 1; cnt = 0
            room_state[dt] = [idx, cnt]
        phong_col.append(int(rooms[idx][0]))
        stt_phong_col.append(cnt + 1)
        room_state[dt][1] = cnt + 1

    df["Phòng thi"] = phong_col
    df["STT trong phòng"] = stt_phong_col

    def make_sbd(row):
        try:
            return f"{int(row['Điểm thi']):02d}{int(row['Phòng thi']):02d}{int(row['STT trong phòng']):02d}"
        except:
            return None
    df["SBD"] = df.apply(make_sbd, axis=1)

    # ── Sort Phase 2 ──
    df = df.sort_values(
        ["Điểm thi","Phòng thi","STT trong phòng"],
        ascending=True, na_position="last"
    ).reset_index(drop=True)
    df["STT"] = range(1, len(df) + 1)

    final_cols = ["STT","Điểm thi","Phòng thi","STT trong phòng","SBD",
                  "ID gốc","Họ và tên đệm","Tên","Ngày sinh","Tháng sinh",
                  "Năm sinh","Giới tính","Khối lớp","Lớp","Xã/Phường",
                  "Tỉnh thành","Cấp độ","HS nước ngoài"]
    return df[[c for c in final_cols if c in df.columns]]


def export_excel(df_sbd, df_diem_raw, df_xep_raw):
    wb = openpyxl.Workbook()
    hdr_fill = PatternFill("solid", fgColor="1D4ED8")
    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(**{s: Side(style="thin", color="D1D5DB") for s in ["left","right","top","bottom"]})
    sbd_fill = PatternFill("solid", fgColor="EFF6FF")
    sbd_font = Font(bold=True, color="1D4ED8", name="Arial Narrow", size=11)

    def write_sheet(ws, df):
        ws.freeze_panes = "A2"
        for ci, col in enumerate(df.columns, 1):
            c = ws.cell(1, ci, col)
            c.font = hdr_font; c.fill = hdr_fill; c.alignment = center; c.border = thin
        for ri, row in enumerate(df.itertuples(index=False), 2):
            for ci, val in enumerate(row, 1):
                c = ws.cell(ri, ci, val)
                c.border = thin
                c.alignment = Alignment(vertical="center",
                                        horizontal="center" if ci <= 5 else "left")
                if ci == 5:
                    c.fill = sbd_fill; c.font = sbd_font
        for col in ws.columns:
            mx = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(mx + 3, 30)
        ws.row_dimensions[1].height = 28

    ws_sbd = wb.active
    ws_sbd.title = "SBD"
    write_sheet(ws_sbd, df_sbd)

    ws_dt = wb.create_sheet("Điểm-thi")
    df_dt = df_diem_raw[["Điểm thi","Phòng","Số lượng hs/phòng"]].copy()
    df_dt = df_dt[pd.to_numeric(df_dt["Điểm thi"], errors="coerce").notna()]
    write_sheet(ws_dt, df_dt)

    ws_xep = wb.create_sheet("Xếp-điểm-thi")
    df_x = df_xep_raw[["Cụm","Điểm thi"]].dropna(subset=["Cụm"]).copy()
    df_x = df_x[df_x["Cụm"].astype(str).str.strip() != ""]
    write_sheet(ws_xep, df_x)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─── SESSION STATE ─────────────────────────────────────────────────────────────
for k in ["result_df", "result_bytes", "errors"]:
    if k not in st.session_state:
        st.session_state[k] = None


# ═══════════════════════════════════════════════════════════════════════════════
# UI
# ═══════════════════════════════════════════════════════════════════════════════
st.markdown('<div class="app-title">CÁC KỲ THI: TẠO SBD TỰ ĐỘNG</div>', unsafe_allow_html=True)

st.markdown('<div class="steps-title" style="text-align: center;">BEBRAS - AMC - VEO</div></br></br>', unsafe_allow_html=True)

# ── BƯỚC 1: Chọn kỳ thi ───────────────────────────────────────────────────────
# st.markdown("""
# <div class="row-card">
#   <div class="row-label">Bước 1</div>
#   <div class="row-content" id="step1-content"></div>
# </div>
# """, unsafe_allow_html=True)
st.markdown('<div class="row-divider"></div>', unsafe_allow_html=True)

col_lbl1, col_body1 = st.columns([1, 4], gap="small")
with col_lbl1:
    # st.markdown('<div class="step-badge">Bước 1</div><br>Chọn kỳ thi', unsafe_allow_html=True)
    # st.markdown('<div class="step-badge">Bước 1</div><div class="step-title">Chọn kỳ thi</div>', unsafe_allow_html=True)
    # st.markdown('<div class="step-badge" style="text-align: center;">Bước 1</div><div class="step-title" style="text-align: center;">Chọn kỳ thi</div>', unsafe_allow_html=True)
    st.markdown('''
<div style="text-align: center;">
    <div class="step-badge">Bước 1</div>
    <div class="step-title">Chọn kỳ thi</div>
</div>
''', unsafe_allow_html=True)
with col_body1:
    exam_key = st.radio(
        "Chọn kỳ thi",
        options=list(EXAM_RULES.keys()),
        format_func=lambda k: EXAM_RULES[k]["label"],
        horizontal=True,
        label_visibility="collapsed",
    )
    # Note cho từng kỳ thi
    EXAM_NOTES = {
        "BEBRAS":  "Cấp độ 1: Lớp 1–2 · Cấp độ 2: Lớp 3–4 · Cấp độ 3: Lớp 5–6 · Cấp độ 4: Lớp 7–8 · Cấp độ 5: Lớp 9–10 · Cấp độ 6: Lớp 11–12",
        "AMC8":    "Dành cho học sinh Lớp 4 – 8",
        "AMC1012": "AMC10: Lớp 6–10 · AMC12: Lớp 11–12",
        "VEO":     "VEO JUNIOR: Lớp 6–9 · VEO: Lớp 10–12",
    }
    st.markdown(f'<div class="exam-note">ℹ️ {EXAM_NOTES[exam_key]}</div>', unsafe_allow_html=True)

st.markdown('<div class="row-divider"></div>', unsafe_allow_html=True)

# ── BƯỚC 2: File dữ liệu ──────────────────────────────────────────────────────
col_lbl2, col_body2 = st.columns([1, 4], gap="small")
with col_lbl2:
    # st.markdown('<div class="step-badge">Bước 2</div><br>Upload', unsafe_allow_html=True)
    st.markdown('''
<div style="text-align: center;">
    <div class="step-badge">Bước 2</div>
    <div class="step-title">Upload</div>
</div>
''', unsafe_allow_html=True)
with col_body2:
    # st.markdown('<div class="sub-label">📥 Tải file mẫu</div>', unsafe_allow_html=True)
    st.download_button(
        "⬇️ Tải file mẫu",
        data=load_sample_bytes(),
        file_name="SBD_mẫu.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    # st.markdown('<div class="sub-label" style="margin-top:.9rem;">📤 Upload file data</div>', unsafe_allow_html=True)
    uploaded = st.file_uploader(
        "Upload file .xlsx",
        type=["xlsx"],
        label_visibility="collapsed",
    )

st.markdown('<div class="row-divider"></div>', unsafe_allow_html=True)

# ── BƯỚC 3: Xử lý ─────────────────────────────────────────────────────────────
col_lbl3, col_body3 = st.columns([1, 4], gap="small")
with col_lbl3:
    # st.markdown('<div class="step-badge">Bước 3</div><br>Xử lý', unsafe_allow_html=True)
    st.markdown('''
<div style="text-align: center;">
    <div class="step-badge">Bước 3</div>
    <div class="step-title">Xử lý</div>
</div>
''', unsafe_allow_html=True)
with col_body3:
    if uploaded:
        if st.button("🚀 Tạo SBD", type="primary", use_container_width=False):
            st.session_state.result_df = None
            st.session_state.result_bytes = None
            st.session_state.errors = None
            with st.spinner("Đang xử lý..."):
                try:
                    xl = pd.ExcelFile(uploaded)
                    missing_sheets = [s for s in ["Học-sinh","Điểm-thi","Xếp-điểm-thi"] if s not in xl.sheet_names]
                    if missing_sheets:
                        st.session_state.errors = [f"File thiếu sheet: **{', '.join(missing_sheets)}**"]
                    else:
                        def read(s):
                            df = pd.read_excel(uploaded, sheet_name=s)
                            return df[[c for c in df.columns if not str(c).startswith("Unnamed")]]
                        df_hs   = read("Học-sinh")
                        df_diem = read("Điểm-thi")
                        df_xep  = read("Xếp-điểm-thi")
                        errs = validate_input(df_hs, df_diem, df_xep, exam_key)
                        if errs:
                            st.session_state.errors = errs
                        else:
                            df_res = process_sbd(df_hs, df_diem, df_xep, exam_key)
                            st.session_state.result_df = df_res
                            st.session_state.result_bytes = export_excel(df_res, df_diem, df_xep)
                except Exception as e:
                    st.session_state.errors = [f"Lỗi: {e}", f"```\n{traceback.format_exc()}\n```"]
    else:
        st.markdown('<span style="color:#94A3B8;font-size:.9rem;">← Upload file ở Bước 2 trước</span>', unsafe_allow_html=True)

st.markdown('<div class="row-divider"></div>', unsafe_allow_html=True)
st.markdown('</div>', unsafe_allow_html=True)  # close steps-wrapper

# ── Lỗi + sửa trực tiếp ──────────────────────────────────────────────────────
if st.session_state.errors:
    for err in st.session_state.errors:
        st.markdown(f'<div class="banner-err">⚠️ {err}</div>', unsafe_allow_html=True)
    if uploaded:
        st.markdown("**🛠️ Chỉnh sửa trực tiếp để khắc phục:**")
        try:
            xl2 = pd.ExcelFile(uploaded)
            sheet_fix = st.selectbox("Sheet cần sửa", xl2.sheet_names)
            df_fix = pd.read_excel(uploaded, sheet_name=sheet_fix)
            df_fix = df_fix[[c for c in df_fix.columns if not str(c).startswith("Unnamed")]]
            edited = st.data_editor(df_fix, num_rows="dynamic", use_container_width=True, key=f"edit_{sheet_fix}")
            if st.button("🔄 Xử lý lại"):
                try:
                    all_sheets = {}
                    for s in xl2.sheet_names:
                        tmp = pd.read_excel(uploaded, sheet_name=s)
                        all_sheets[s] = tmp[[c for c in tmp.columns if not str(c).startswith("Unnamed")]]
                    all_sheets[sheet_fix] = edited
                    df_hs2   = all_sheets.get("Học-sinh", pd.DataFrame())
                    df_diem2 = all_sheets.get("Điểm-thi", pd.DataFrame())
                    df_xep2  = all_sheets.get("Xếp-điểm-thi", pd.DataFrame())
                    errs2 = validate_input(df_hs2, df_diem2, df_xep2, exam_key)
                    if errs2:
                        st.session_state.errors = errs2
                        st.rerun()
                    else:
                        df_res2 = process_sbd(df_hs2, df_diem2, df_xep2, exam_key)
                        st.session_state.result_df = df_res2
                        st.session_state.result_bytes = export_excel(df_res2, df_diem2, df_xep2)
                        st.session_state.errors = None
                        st.rerun()
                except Exception as e2:
                    st.error(f"Vẫn còn lỗi: {e2}")
        except Exception:
            pass

# ── Kết quả ──────────────────────────────────────────────────────────────────
if st.session_state.result_df is not None:
    df_res = st.session_state.result_df
    st.markdown(f'<div class="banner-ok">✅ Xử lý thành công — {len(df_res):,} học sinh.</div>', unsafe_allow_html=True)
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Tổng HS", f"{len(df_res):,}")
    c2.metric("Điểm thi", df_res["Điểm thi"].nunique() if "Điểm thi" in df_res else "—")
    c3.metric("Phòng thi", df_res.groupby(["Điểm thi","Phòng thi"]).ngroups if "Phòng thi" in df_res else "—")
    c4.metric("Cấp độ", df_res["Cấp độ"].nunique() if "Cấp độ" in df_res else "—")
    st.dataframe(df_res.head(100), use_container_width=True, height=320)
    if len(df_res) > 100:
        st.caption(f"Hiển thị 100 / {len(df_res):,} dòng.")
    st.download_button(
        "📥 Tải file kết quả (.xlsx)",
        data=st.session_state.result_bytes,
        file_name=f"SBD_{exam_key}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )

# st.markdown(
#     '<p style="text-align:center;color:#94A3B8;font-size:.78rem;margin-top:1.5rem;">'
#     'SBD Auto Generator · BEBRAS · AMC8 · AMC10/12 · VEO</p>',
#     unsafe_allow_html=True,
# )
